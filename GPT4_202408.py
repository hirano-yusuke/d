import openpyxl
import pandas as pd
import unicodedata
import re
from openai import AzureOpenAI

# ExcelファイルのD列(各人メモ記載列)を取得
def extract_column_data(file_path, column_letter='D'):
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook.active
    return [cell.value for cell in sheet[column_letter]]

# ExcelファイルのC列(質問記載列)を取得
def extract_column_Question(file_path, column_letter='C'):
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook.active
    return [cell.value for cell in sheet[column_letter]]

# 結果ファイルのE列から前年度の回答を取得する
def extract_column_Ans(file_path, column_letter='E'):
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook.active
    return [cell.value for cell in sheet[column_letter]]

# メモ内の指摘事項の重要度を抽出
def count_target_words_filter_positive(texts, target_words):
    word_counts = {word: 0 for word in target_words}
    
    for text in texts:
        for word in target_words:
            word_counts[word] += text.count(word)
    
    # 出現していない重要度の単語は出力しない
    filtered_counts = {word: count for word, count in word_counts.items() if count > 0}
    
    return filtered_counts

# GPTクライアント作成
client = AzureOpenAI(
    api_version="2024-02-15-preview",
    #GPT4用
    api_key="3aed0476a1d84d91b6c61d1a475b7046",
    azure_endpoint="https://hirano-gpt4.openai.azure.com/"
#     #GPT3.5用
#     api_key="2b26e729e3cf4ee2b483a1953b333bc0",
#     azure_endpoint="https://denka-openai-jaeast.openai.azure.com/"
 )

# ファイルパスの設定
files1 = ['C:\\Python\\MVP\\01_work\\docs\\03-gptapp\\docs01\\memberA.xlsx']
files2 = ['C:\\Python\\MVP\\01_work\\docs\\03-gptapp\\docs02\\memberB.xlsx']
files3 = ['C:\\Python\\MVP\\01_work\\docs\\03-gptapp\\docs03\\memberC.xlsx']
files4 = ['C:\\Python\\MVP\\01_work\\docs\\03-gptapp\\docs04\\memberD.xlsx']
example = ['C:\\Python\\MVP\\01_work\\docs\\03-gptapp\\Ans\\Ans.xlsx']

# データの取得
textsA = [extract_column_data(file) for file in files1]
textsB = [extract_column_data(file) for file in files2]
textsC = [extract_column_data(file) for file in files3]
textsD = [extract_column_data(file) for file in files4]

# フラット化したリストを作成
textsA_flat = [item for sublist in textsA for item in sublist]
textsB_flat = [item for sublist in textsB for item in sublist]
textsC_flat = [item for sublist in textsC for item in sublist]
textsD_flat = [item for sublist in textsD for item in sublist]

questions = [extract_column_Question(file) for file in files1]
questions_flat = [item for sublist in questions for item in sublist]

Ans = [extract_column_Ans(file) for file in example]
Ans_flat = [item for sublist in Ans for item in sublist]

# 要約の作成関数
def summarize_text(client, notes, q_element, length):
    response = client.chat.completions.create(
        model="GPT4-TEST",  #GPTモデルの内容"gpt35turbo16k","AzureのGPTデプロイ名であるGPT4-TEST"
        temperature=0.2,
        messages=[
            {"role": "system", "content": "You are an outstanding manager of an audit department. Please summarize the following text in Japanese."},
            {"role": "user", "content": f"The sentence you want summarized is {notes} Please consider that this input is the hearing content regarding question {q_element}."},
            {"role": "user", "content": f"Please summarize in Japanese to a maximum of {length} characters. Please delete any English sentences in the summary."}
        ]
    )
    return response.choices[0].message.content

# 要約の生成と重要度の抽出
results_200 = []
results_300 = []
results_400 = []
classifications = []

for textsA, textsB, textsC, textsD, q_element in zip(textsA_flat, textsB_flat, textsC_flat, textsD_flat, questions_flat):
    if textsA is None and textsB is None and textsC is None and textsD is None:
        results_200.append(None)
        results_300.append(None)
        results_400.append(None)
        classifications.append(None)
    else:
        notes = [note for note in [textsA, textsB, textsC, textsD] if note is not None]
        notes_str = ' '.join(notes)
        
        # 要約の生成
        results_200.append(summarize_text(client, notes_str, q_element, 200))
        results_300.append(summarize_text(client, notes_str, q_element, 300))
        results_400.append(summarize_text(client, notes_str, q_element, 400))
        
        # メモ内の重要度に関する言及の抽出
        target_words_example = ["CAR", "OBS", "OFI", "メモ"]
        notes_half_width = [unicodedata.normalize('NFKC', s) for s in notes]
        result_filtered = count_target_words_filter_positive(notes_half_width, target_words_example)
        classifications.append(result_filtered)

# DataFrameの作成
df_output = pd.DataFrame({
    "Summary_200": results_200,
    "Summary_300": results_300,
    "Summary_400": results_400,
    "Classification": classifications
})

# CSVファイルに出力
df_output.to_csv('summary_outputGPT4.csv', index=False)
